# Constants
DIR_PATH = r'C:\Users\appelm3\OneDrive - Medtronic PLC\Survey_Temp'
OUTPUT_FILE = 'output.xlsx'

# pip install pandas openpyxl matplotlib extract-msg

import os
import re
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import extract_msg


def get_email_message(file_path):
    msg = extract_msg.Message(file_path)
    mail_body = msg.body
    return mail_body


def extract_data_from_mail(mail_body):
    # Extract basic details
    id_info_patterns = {
        "Date": r"Date: (\d{2}.\d{2}.\d{4})",
        "Time": r"Time: (\d{2}:\d{2}:\d{2})",
        "Reader Name": r"Reader name: ([^\n]*)",
        "Study Name": r"Study name: ([^\n]*)",
        "Session Time": r"Session time: ([^\n]+)"
    }
    
    extracted_info = {}
    for key, pattern in id_info_patterns.items():
        match = re.search(pattern, mail_body)
        extracted_info[key] = match.group(1) if match else ''


  #  extracted_info = {
  #      key: re.search(pattern, mail_body).group(1) if (match := re.search(pattern, mail_body)) else ''
  #      for key, pattern in id_info_patterns.items()
  #  }
  
  

    # Create DataFrame for basic details
    details = [extracted_info['Reader Name'], extracted_info['Study Name'], extracted_info['Session Time']]
    details_df = pd.DataFrame([details], columns=['Reader Name', 'Study Name', 'Session Time'])

    # Extract survey responses
    results_pattern = r'Question: (.*?)\nAnswer: (.*?)\nComments: (.*?)\n'
    matches = re.findall(results_pattern, mail_body, re.DOTALL)

    # Organize survey responses
    data = {}
    questions = [
        "In comparison to the current PillCam video",
        "If the video review was AI-assisted, how would you rate the AI experience?",
        "If the video review was AI-assisted, how would you rate the AI user-interface bounding-boxes presentation?",
        "In case the AI assisted reading followed the reading of the same case without AI assistance, did the AI increase your confidence in the clinical diagnosis and/or assist with your interpretation?"
    ]

    for match in matches:
        question, answer, comments = map(str.strip, match)

        for i, q in enumerate(questions, start=1):
            if question.startswith(q):
                data[f'Answer {i}'] = answer
                data[f'Comments {i}'] = comments

    # Create DataFrame for survey responses
    ans_df = pd.DataFrame([data])

    # Combine basic details and survey responses
    df = pd.concat([details_df, ans_df], axis=1)
    return df


def create_bar_plot(header, label_counts, question, image_row):
    max_count = label_counts.max()
    total_count = label_counts.sum()
    percentages = (label_counts / total_count * 100).round(2)

    # Plotting
    plt.figure(figsize=(9, 6))
    ax = plt.gca()
    ax.bar(label_counts.index, label_counts, width=0.8, color='skyblue')
    plt.title(question)
    plt.xticks(rotation=45, ha="right")
    plt.ylabel('Count'), plt.ylim([0, max_count + 2])
    plt.yticks(range(0, int(max_count) + 2, 2))

    # Add annotations
    for i, (count, percentage) in enumerate(zip(label_counts, percentages)):
        ax.text(i, count, f'{count} ({percentage}%)', ha='center', va='bottom')

    # Additional details
    plt.text(-0.5, max_count, 'Total: ' + str(total_count))
    plt.tight_layout()

    # Save the plot and close the figure
    plot_image = f'{header}_bar_plot.png'
    plt.savefig(plot_image)
    plt.close()

    # Add the plot image to the Excel sheet
    img = Image(plot_image)
    img.anchor = f'A{image_row}'
    return img


def export_plots_to_excel(df):
    # Define labels and titles
    labels = {
        "Answer 1": ["Significantly shorter", "Slightly shorter", "Pretty much the same", "Slightly longer",
                     "Significantly longer"],
        "Answer 2": ["Burdensome, mostly annoying false alarms", "Some false alarms, overall ok",
                     "Excellent, very helpful!", "A few misses here and there, overall ok",
                     "Many misdetections of significant lesions"],
        "Answer 3": ["Did not like at all", "Can be improved, overall ok", "Clear and user friendly",
                     "Excellent! Very helpful"],
        "Answer 4": ["No", "Unsure", "Yes"]
    }

    Titles = [
        'Rate of the new SB Genius video, compared to the current PillCam video',
        'Rate of the AI experience',
        'Rate of the bounding-boxes presentation',
        'Did the AI increase your confidence in the clinical diagnosis and/or assist with your interpretation?'
    ]

    # Load or create the workbook and sheet
    try:
        wb = load_workbook("output.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    ws_name = "sheet2"
    if ws_name in wb.sheetnames:
        ws = wb[ws_name]
    else:
        ws = wb.create_sheet(ws_name)

    image_row = 2

    # Generate and add images to the sheet
    for i, (header, label) in enumerate(labels.items()):
        data_series = df[header]
        label_counts = pd.Series([data_series.str.contains(lbl).sum() for lbl in label], index=label)

        img = create_bar_plot(header, label_counts, Titles[i], image_row)
        ws.add_image(img)

        image_row += 30

    # Save the workbook
    wb.save("output.xlsx")


def main():
    survey_data = pd.DataFrame()

    for message in os.listdir(DIR_PATH):
        mail_body = get_email_message(os.path.join(DIR_PATH, message))
        new_data = extract_data_from_mail(mail_body)
        survey_data = pd.concat([survey_data, new_data])

    survey_data.to_excel(OUTPUT_FILE, index=False)
    export_plots_to_excel(survey_data)


if __name__ == "__main__":
    main()